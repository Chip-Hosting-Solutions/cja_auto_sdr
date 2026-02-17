"""Tests covering uncovered lines in diff output writers and helpers in generator.py.

Targets:
- _format_diff_value except TypeError/ValueError (lines 3262-3263)
- write_diff_console_output summary_only with no changes (line 3107)
- _format_side_by_side no changed_fields early return (line 3316)
- write_diff_grouped_by_field_output breaking changes, added/removed with limit
  (lines 3400, 3413-3418, 3446-3451)
- detect_breaking_changes schemaPath path (line 3627)
- write_diff_json_output exception handler (lines 3784-3786)
- write_diff_markdown_output edge cases + exception handler (lines 3862, 3888, 3913, 3961-3963)
- write_diff_html_output edge cases + exception handler (lines 4248, 4256, 4260-4261,
  4297, 4300, 4358-4360)
- write_diff_excel_output no-changes sheet + exception handler (lines 4484-4486, 4522,
  4569-4571)
- write_diff_csv_output exception handler (lines 4707, 4734-4736)
- write_diff_output group_by_field routing (lines 4777-4780)
"""

from __future__ import annotations

import logging
from unittest.mock import patch

import pytest

from cja_auto_sdr.diff.models import (
    ChangeType,
    ComponentDiff,
    DiffResult,
    DiffSummary,
    InventoryItemDiff,
    MetadataDiff,
)
from cja_auto_sdr.generator import (
    _format_diff_value,
    _format_side_by_side,
    detect_breaking_changes,
    write_diff_console_output,
    write_diff_csv_output,
    write_diff_excel_output,
    write_diff_grouped_by_field_output,
    write_diff_html_output,
    write_diff_json_output,
    write_diff_markdown_output,
    write_diff_output,
)

# ==================== Helpers ====================


def _make_diff_result(
    metric_diffs=None,
    dimension_diffs=None,
    has_changes=True,
    summary_kwargs=None,
    calc_metrics_diffs=None,
    segments_diffs=None,
):
    """Build a minimal DiffResult with sensible defaults."""
    summary_kw = {
        "source_metrics_count": 5,
        "target_metrics_count": 5,
        "source_dimensions_count": 3,
        "target_dimensions_count": 3,
        "metrics_modified": 1 if has_changes else 0,
    }
    if summary_kwargs:
        summary_kw.update(summary_kwargs)
    return DiffResult(
        summary=DiffSummary(**summary_kw),
        metadata_diff=MetadataDiff(
            source_name="Source DV",
            target_name="Target DV",
            source_id="dv_src",
            target_id="dv_tgt",
        ),
        metric_diffs=metric_diffs or [],
        dimension_diffs=dimension_diffs or [],
        calc_metrics_diffs=calc_metrics_diffs,
        segments_diffs=segments_diffs,
    )


def _make_logger():
    """Return a logger for tests."""
    return logging.getLogger("test")


# ==================== TestFormatDiffValue ====================


class TestFormatDiffValue:
    """Tests for _format_diff_value handling of values that make pd.isna raise."""

    def test_list_triggers_except_path(self):
        """A list causes pd.isna() to return an array; the if-check raises ValueError."""
        result = _format_diff_value([1, 2, 3])
        assert result == "[1, 2, 3]"

    def test_dict_triggers_except_path(self):
        """A dict causes pd.isna() to return a non-scalar; the if-check raises."""
        result = _format_diff_value({"key": "value"})
        assert result == "{'key': 'value'}"

    def test_list_with_truncation(self):
        """List value that exceeds max_len is truncated."""
        long_list = list(range(50))
        result = _format_diff_value(long_list, truncate=True, max_len=20)
        assert len(result) == 20

    def test_list_no_truncation(self):
        """List value without truncation returns full string."""
        val = [10, 20, 30]
        result = _format_diff_value(val, truncate=False)
        assert result == "[10, 20, 30]"


# ==================== TestDiffConsoleOutputSummaryOnly ====================


class TestDiffConsoleOutputSummaryOnly:
    """Test write_diff_console_output with summary_only=True and no changes."""

    def test_summary_only_no_changes(self):
        """Line 3107: summary_only path prints 'No differences found.' when no changes."""
        dr = _make_diff_result(has_changes=False)
        output = write_diff_console_output(dr, summary_only=True, use_color=False)
        assert "No differences found." in output

    def test_summary_only_no_changes_with_color(self):
        """summary_only path with color enabled still contains the message."""
        dr = _make_diff_result(has_changes=False)
        output = write_diff_console_output(dr, summary_only=True, use_color=True)
        assert "No differences found." in output

    def test_summary_only_with_changes(self):
        """summary_only path with changes shows total changes instead."""
        dr = _make_diff_result(
            has_changes=True,
            metric_diffs=[
                ComponentDiff(
                    id="m1",
                    name="Metric1",
                    change_type=ChangeType.MODIFIED,
                    changed_fields={"desc": ("old", "new")},
                ),
            ],
        )
        output = write_diff_console_output(dr, summary_only=True, use_color=False)
        assert "Total changes:" in output


# ==================== TestFormatSideBySideEarlyReturn ====================


class TestFormatSideBySideEarlyReturn:
    """Test _format_side_by_side returns empty list for non-modified or empty fields."""

    def test_added_component_returns_empty(self):
        """Line 3316: ADDED component returns [] immediately."""
        diff = ComponentDiff(id="m1", name="Met", change_type=ChangeType.ADDED)
        assert _format_side_by_side(diff, "Src", "Tgt") == []

    def test_removed_component_returns_empty(self):
        """REMOVED component returns [] immediately."""
        diff = ComponentDiff(id="m1", name="Met", change_type=ChangeType.REMOVED)
        assert _format_side_by_side(diff, "Src", "Tgt") == []

    def test_modified_no_changed_fields_returns_empty(self):
        """MODIFIED component with empty changed_fields returns []."""
        diff = ComponentDiff(id="m1", name="Met", change_type=ChangeType.MODIFIED, changed_fields={})
        assert _format_side_by_side(diff, "Src", "Tgt") == []

    def test_unchanged_returns_empty(self):
        """UNCHANGED component returns [] immediately."""
        diff = ComponentDiff(id="m1", name="Met", change_type=ChangeType.UNCHANGED)
        assert _format_side_by_side(diff, "Src", "Tgt") == []


# ==================== TestDiffGroupedByFieldOutput ====================


class TestDiffGroupedByFieldOutput:
    """Test write_diff_grouped_by_field_output breaking changes, added/removed with limit."""

    def _make_breaking_result(self):
        """DiffResult with type and schemaPath changes (breaking changes)."""
        return _make_diff_result(
            metric_diffs=[
                ComponentDiff(
                    id="m1",
                    name="Revenue",
                    change_type=ChangeType.MODIFIED,
                    changed_fields={
                        "type": ("int", "decimal"),
                        "schemaPath": ("_xdm.old.path", "_xdm.new.path"),
                    },
                ),
            ],
            dimension_diffs=[
                ComponentDiff(
                    id="d1",
                    name="Page",
                    change_type=ChangeType.MODIFIED,
                    changed_fields={"type": ("string", "int")},
                ),
            ],
            summary_kwargs={
                "metrics_modified": 1,
                "dimensions_modified": 1,
            },
        )

    def test_breaking_changes_header_appears(self):
        """Lines 3413-3418: Breaking changes section is rendered."""
        dr = self._make_breaking_result()
        output = write_diff_grouped_by_field_output(dr, use_color=False)
        assert "BREAKING CHANGES DETECTED" in output

    def test_breaking_changes_list_type_and_schema(self):
        """Lines 3399-3400: Both type and schemaPath changes are listed."""
        dr = self._make_breaking_result()
        output = write_diff_grouped_by_field_output(dr, use_color=False)
        assert "m1: type changed" in output
        assert "m1: schemaPath changed" in output
        assert "d1: type changed" in output

    def test_breaking_changes_show_old_new_values(self):
        """Lines 3417-3418: Old and new values are displayed."""
        dr = self._make_breaking_result()
        output = write_diff_grouped_by_field_output(dr, use_color=False)
        assert "'int'" in output
        assert "'decimal'" in output

    def test_added_section_with_limit(self):
        """Lines 3446-3451: Added section with limit shows '... and N more'."""
        added_diffs = [ComponentDiff(id=f"m{i}", name=f"Met{i}", change_type=ChangeType.ADDED) for i in range(20)]
        dr = _make_diff_result(
            metric_diffs=added_diffs,
            summary_kwargs={"metrics_added": 20},
        )
        output = write_diff_grouped_by_field_output(dr, use_color=False, limit=5)
        assert "ADDED (20)" in output
        assert "... and 15 more" in output

    def test_removed_section_with_limit(self):
        """Removed section with limit shows '... and N more'."""
        removed_diffs = [ComponentDiff(id=f"d{i}", name=f"Dim{i}", change_type=ChangeType.REMOVED) for i in range(12)]
        dr = _make_diff_result(
            dimension_diffs=removed_diffs,
            summary_kwargs={"dimensions_removed": 12},
        )
        output = write_diff_grouped_by_field_output(dr, use_color=False, limit=3)
        assert "REMOVED (12)" in output
        assert "... and 9 more" in output

    def test_field_changes_with_limit(self):
        """Field change section with limit truncates listing."""
        diffs = [
            ComponentDiff(
                id=f"m{i}",
                name=f"Met{i}",
                change_type=ChangeType.MODIFIED,
                changed_fields={"description": (f"old{i}", f"new{i}")},
            )
            for i in range(15)
        ]
        dr = _make_diff_result(
            metric_diffs=diffs,
            summary_kwargs={"metrics_modified": 15},
        )
        output = write_diff_grouped_by_field_output(dr, use_color=False, limit=5)
        assert "... and 10 more" in output

    def test_unlimited_shows_all(self):
        """limit=0 shows all items without truncation."""
        diffs = [
            ComponentDiff(
                id=f"m{i}",
                name=f"Met{i}",
                change_type=ChangeType.MODIFIED,
                changed_fields={"description": (f"old{i}", f"new{i}")},
            )
            for i in range(15)
        ]
        dr = _make_diff_result(
            metric_diffs=diffs,
            summary_kwargs={"metrics_modified": 15},
        )
        output = write_diff_grouped_by_field_output(dr, use_color=False, limit=0)
        assert "... and" not in output
        for i in range(15):
            assert f"m{i}" in output


# ==================== TestDetectBreakingChangesSchemaPath ====================


class TestDetectBreakingChangesSchemaPath:
    """Test detect_breaking_changes for schemaPath changes (line 3627)."""

    def test_schema_path_change_detected(self):
        """Line 3627: schemaPath change is flagged as breaking with severity 'medium'."""
        dr = _make_diff_result(
            metric_diffs=[
                ComponentDiff(
                    id="m1",
                    name="Revenue",
                    change_type=ChangeType.MODIFIED,
                    changed_fields={"schemaPath": ("_xdm.old.path", "_xdm.new.path")},
                ),
            ],
        )
        breaking = detect_breaking_changes(dr)
        assert len(breaking) == 1
        assert breaking[0]["change_type"] == "schema_changed"
        assert breaking[0]["severity"] == "medium"
        assert breaking[0]["field"] == "schemaPath"

    def test_both_type_and_schema_detected(self):
        """Both type and schemaPath changes produce separate breaking entries."""
        dr = _make_diff_result(
            metric_diffs=[
                ComponentDiff(
                    id="m1",
                    name="Revenue",
                    change_type=ChangeType.MODIFIED,
                    changed_fields={
                        "type": ("int", "decimal"),
                        "schemaPath": ("_xdm.old.path", "_xdm.new.path"),
                    },
                ),
            ],
        )
        breaking = detect_breaking_changes(dr)
        assert len(breaking) == 2
        types_found = {b["change_type"] for b in breaking}
        assert types_found == {"type_changed", "schema_changed"}


# ==================== TestWriteDiffJsonOutputErrors ====================


class TestWriteDiffJsonOutputErrors:
    """Test write_diff_json_output generic exception handler (lines 3784-3786)."""

    def test_runtime_error_propagates(self, tmp_path):
        """Mock builtins.open to raise RuntimeError; verify it re-raises."""
        dr = _make_diff_result()
        logger = _make_logger()
        with patch("builtins.open", side_effect=RuntimeError("disk full")):
            with pytest.raises(RuntimeError, match="disk full"):
                write_diff_json_output(dr, "diff_report", str(tmp_path), logger)


# ==================== TestWriteDiffMarkdownOutputEdges ====================


class TestWriteDiffMarkdownOutputEdges:
    """Test write_diff_markdown_output edge cases and exception handler."""

    def test_no_changes_message(self, tmp_path):
        """Line 3862: 'No differences found.' appears when no changes."""
        dr = _make_diff_result(has_changes=False)
        logger = _make_logger()
        path = write_diff_markdown_output(dr, "diff_report", str(tmp_path), logger)
        with open(path) as f:
            content = f.read()
        assert "No differences found." in content

    def test_no_metric_changes_shows_no_changes(self, tmp_path):
        """Line 3888: '*No changes*' appears for metrics section with no changes."""
        dr = _make_diff_result(
            has_changes=False,
            metric_diffs=[],
            dimension_diffs=[],
        )
        logger = _make_logger()
        path = write_diff_markdown_output(dr, "diff_report", str(tmp_path), logger, changes_only=False)
        with open(path) as f:
            content = f.read()
        assert "*No changes*" in content

    def test_no_dimension_changes_shows_no_changes(self, tmp_path):
        """Line 3913: '*No changes*' for dimensions when no dim changes, changes_only=False."""
        dr = _make_diff_result(
            metric_diffs=[
                ComponentDiff(id="m1", name="Met", change_type=ChangeType.ADDED),
            ],
            dimension_diffs=[],
            summary_kwargs={"metrics_added": 1, "metrics_modified": 0},
        )
        logger = _make_logger()
        path = write_diff_markdown_output(dr, "diff_report", str(tmp_path), logger, changes_only=False)
        with open(path) as f:
            content = f.read()
        # Metrics section should have content, dimensions section should say *No changes*
        assert "## Dimensions Changes" in content
        assert "*No changes*" in content

    def test_side_by_side_mode(self, tmp_path):
        """Side-by-side tables appear for modified items in markdown."""
        dr = _make_diff_result(
            metric_diffs=[
                ComponentDiff(
                    id="m1",
                    name="Revenue",
                    change_type=ChangeType.MODIFIED,
                    changed_fields={"description": ("old desc", "new desc")},
                ),
            ],
        )
        logger = _make_logger()
        path = write_diff_markdown_output(
            dr,
            "diff_report",
            str(tmp_path),
            logger,
            side_by_side=True,
        )
        with open(path) as f:
            content = f.read()
        assert "Side by Side" in content
        assert "`m1`" in content

    def test_side_by_side_with_long_values(self, tmp_path):
        """Lines 4003-4006: Long values in side-by-side markdown are truncated at 50 chars."""
        long_value = "x" * 100
        dr = _make_diff_result(
            metric_diffs=[
                ComponentDiff(
                    id="m1",
                    name="Revenue",
                    change_type=ChangeType.MODIFIED,
                    changed_fields={"description": (long_value, "short")},
                ),
            ],
        )
        logger = _make_logger()
        path = write_diff_markdown_output(
            dr,
            "diff_report",
            str(tmp_path),
            logger,
            side_by_side=True,
        )
        with open(path) as f:
            content = f.read()
        # Value should be truncated with "..."
        assert "..." in content

    def test_inventory_diffs_in_markdown(self, tmp_path):
        """Inventory sections (calc metrics, segments) appear in markdown."""
        dr = _make_diff_result(
            calc_metrics_diffs=[
                InventoryItemDiff(
                    id="cm1",
                    name="CalcMet1",
                    change_type=ChangeType.ADDED,
                    inventory_type="calculated_metric",
                ),
            ],
            segments_diffs=[
                InventoryItemDiff(
                    id="s1",
                    name="Seg1",
                    change_type=ChangeType.REMOVED,
                    inventory_type="segment",
                ),
            ],
            summary_kwargs={
                "source_calc_metrics_count": 1,
                "target_calc_metrics_count": 2,
                "calc_metrics_added": 1,
                "source_segments_count": 2,
                "target_segments_count": 1,
                "segments_removed": 1,
            },
        )
        logger = _make_logger()
        path = write_diff_markdown_output(dr, "diff_report", str(tmp_path), logger)
        with open(path) as f:
            content = f.read()
        assert "Calculated Metrics Changes" in content
        assert "Segments Changes" in content

    def test_exception_handler(self, tmp_path):
        """Lines 3961-3963: Generic exception re-raises."""
        dr = _make_diff_result()
        logger = _make_logger()
        with patch("builtins.open", side_effect=RuntimeError("write failed")):
            with pytest.raises(RuntimeError, match="write failed"):
                write_diff_markdown_output(dr, "diff_report", str(tmp_path), logger)


# ==================== TestWriteDiffHtmlOutputEdges ====================


class TestWriteDiffHtmlOutputEdges:
    """Test write_diff_html_output edge cases and exception handler."""

    def test_no_changes_html(self, tmp_path):
        """Line 4248: 'No differences found.' message in HTML."""
        dr = _make_diff_result(has_changes=False)
        logger = _make_logger()
        path = write_diff_html_output(dr, "diff_report", str(tmp_path), logger)
        with open(path) as f:
            content = f.read()
        assert "No differences found." in content

    def test_changes_only_empty_diffs(self, tmp_path):
        """Lines 4255-4256: changes_only=True with no changes returns empty for section."""
        dr = _make_diff_result(
            has_changes=False,
            metric_diffs=[],
            dimension_diffs=[],
        )
        logger = _make_logger()
        path = write_diff_html_output(
            dr,
            "diff_report",
            str(tmp_path),
            logger,
            changes_only=True,
        )
        with open(path) as f:
            content = f.read()
        # With changes_only=True and no changes, diff tables should be empty strings
        assert "No differences found." in content

    def test_no_changes_in_section_shows_no_changes(self, tmp_path):
        """Lines 4260-4261: section with no changes shows '<em>No changes</em>'."""
        dr = _make_diff_result(
            has_changes=False,
            metric_diffs=[],
            dimension_diffs=[],
        )
        logger = _make_logger()
        path = write_diff_html_output(
            dr,
            "diff_report",
            str(tmp_path),
            logger,
            changes_only=False,
        )
        with open(path) as f:
            content = f.read()
        assert "<em>No changes</em>" in content

    def test_inventory_diff_table_none_returns_empty(self, tmp_path):
        """Line 4297: generate_inventory_diff_table with None returns ''."""
        dr = _make_diff_result(
            calc_metrics_diffs=[
                InventoryItemDiff(
                    id="cm1",
                    name="CalcMet1",
                    change_type=ChangeType.ADDED,
                    inventory_type="calculated_metric",
                ),
            ],
            segments_diffs=None,
            summary_kwargs={
                "source_calc_metrics_count": 1,
                "target_calc_metrics_count": 2,
                "calc_metrics_added": 1,
            },
        )
        logger = _make_logger()
        path = write_diff_html_output(dr, "diff_report", str(tmp_path), logger)
        with open(path) as f:
            content = f.read()
        assert "Calculated Metrics Changes" in content
        # segments_diffs=None should not produce a Segments Changes section
        assert "Segments Changes" not in content

    def test_inventory_changes_only_no_changes(self, tmp_path):
        """Line 4300: changes_only=True with unchanged inventory returns empty."""
        dr = _make_diff_result(
            calc_metrics_diffs=[
                InventoryItemDiff(
                    id="cm1",
                    name="CalcMet1",
                    change_type=ChangeType.UNCHANGED,
                    inventory_type="calculated_metric",
                ),
            ],
            segments_diffs=[],
            summary_kwargs={
                "source_calc_metrics_count": 1,
                "target_calc_metrics_count": 1,
            },
        )
        logger = _make_logger()
        path = write_diff_html_output(
            dr,
            "diff_report",
            str(tmp_path),
            logger,
            changes_only=True,
        )
        with open(path) as f:
            content = f.read()
        # With changes_only and only unchanged items, inventory table is empty
        assert "Inventory Changes" in content

    def test_exception_handler(self, tmp_path):
        """Lines 4358-4360: Generic exception re-raises."""
        dr = _make_diff_result()
        logger = _make_logger()
        with patch("builtins.open", side_effect=RuntimeError("html fail")):
            with pytest.raises(RuntimeError, match="html fail"):
                write_diff_html_output(dr, "diff_report", str(tmp_path), logger)


# ==================== TestWriteDiffExcelOutputEdges ====================


class TestWriteDiffExcelOutputEdges:
    """Test write_diff_excel_output no-changes sheet and exception handler."""

    def test_no_changes_placeholder_sheet(self, tmp_path):
        """Lines 4484-4486: Empty diffs produce a 'No changes' placeholder sheet."""
        dr = _make_diff_result(
            has_changes=False,
            metric_diffs=[],
            dimension_diffs=[],
        )
        logger = _make_logger()
        path = write_diff_excel_output(dr, "diff_report", str(tmp_path), logger)
        import openpyxl

        wb = openpyxl.load_workbook(path)
        metrics_ws = wb["Metrics Diff"]
        # First row is header, second row should say "No changes"
        assert metrics_ws.cell(1, 1).value == "Message"
        assert metrics_ws.cell(2, 1).value == "No changes"

    def test_changes_only_empty_produces_placeholder(self, tmp_path):
        """changes_only=True with unchanged items produces placeholder sheet."""
        dr = _make_diff_result(
            metric_diffs=[
                ComponentDiff(id="m1", name="Met", change_type=ChangeType.UNCHANGED),
            ],
            dimension_diffs=[],
            summary_kwargs={"metrics_modified": 0},
        )
        logger = _make_logger()
        path = write_diff_excel_output(
            dr,
            "diff_report",
            str(tmp_path),
            logger,
            changes_only=True,
        )
        import openpyxl

        wb = openpyxl.load_workbook(path)
        metrics_ws = wb["Metrics Diff"]
        assert metrics_ws.cell(1, 1).value == "Message"

    def test_inventory_none_skipped(self, tmp_path):
        """Line 4522: write_inventory_diff_sheet with None diffs returns early."""
        dr = _make_diff_result(
            calc_metrics_diffs=None,
            segments_diffs=None,
        )
        logger = _make_logger()
        path = write_diff_excel_output(dr, "diff_report", str(tmp_path), logger)
        import openpyxl

        wb = openpyxl.load_workbook(path)
        # Should not have inventory sheets
        assert "Calc Metrics Diff" not in wb.sheetnames
        assert "Segments Diff" not in wb.sheetnames

    def test_exception_handler(self, tmp_path):
        """Lines 4569-4571: Generic exception re-raises."""
        dr = _make_diff_result()
        logger = _make_logger()
        with patch("pandas.ExcelWriter", side_effect=RuntimeError("excel fail")):
            with pytest.raises(RuntimeError, match="excel fail"):
                write_diff_excel_output(dr, "diff_report", str(tmp_path), logger)


# ==================== TestWriteDiffCsvOutputEdges ====================


class TestWriteDiffCsvOutputEdges:
    """Test write_diff_csv_output edge cases and exception handler."""

    def test_inventory_none_skipped(self, tmp_path):
        """Line 4707: write_inventory_diff_csv with None diffs returns early."""
        dr = _make_diff_result(
            calc_metrics_diffs=None,
            segments_diffs=None,
        )
        logger = _make_logger()
        import os

        csv_dir = write_diff_csv_output(dr, "diff_report", str(tmp_path), logger)
        files = os.listdir(csv_dir)
        assert "calc_metrics_diff.csv" not in files
        assert "segments_diff.csv" not in files

    def test_exception_handler(self, tmp_path):
        """Lines 4734-4736: Generic exception re-raises."""
        dr = _make_diff_result()
        logger = _make_logger()
        with patch("os.makedirs", side_effect=RuntimeError("csv fail")):
            with pytest.raises(RuntimeError, match="csv fail"):
                write_diff_csv_output(dr, "diff_report", str(tmp_path), logger)


# ==================== TestWriteDiffOutput ====================


class TestWriteDiffOutput:
    """Test write_diff_output group_by_field routing (lines 4777-4780)."""

    def test_group_by_field_console_only(self, tmp_path, capsys):
        """Lines 4777-4780: group_by_field=True with console format routes to grouped output."""
        dr = _make_diff_result(
            metric_diffs=[
                ComponentDiff(
                    id="m1",
                    name="Revenue",
                    change_type=ChangeType.MODIFIED,
                    changed_fields={"description": ("old", "new")},
                ),
            ],
        )
        logger = _make_logger()
        result = write_diff_output(
            dr,
            output_format="console",
            base_filename="diff_test",
            output_dir=str(tmp_path),
            logger=logger,
            group_by_field=True,
            use_color=False,
        )
        captured = capsys.readouterr()
        assert result is not None
        assert "GROUPED BY FIELD" in result
        assert "GROUPED BY FIELD" in captured.out

    def test_group_by_field_all_format(self, tmp_path, capsys):
        """group_by_field=True with 'all' format outputs grouped console plus file formats."""
        dr = _make_diff_result(
            metric_diffs=[
                ComponentDiff(
                    id="m1",
                    name="Revenue",
                    change_type=ChangeType.MODIFIED,
                    changed_fields={"description": ("old", "new")},
                ),
            ],
        )
        logger = _make_logger()
        write_diff_output(
            dr,
            output_format="all",
            base_filename="diff_test",
            output_dir=str(tmp_path),
            logger=logger,
            group_by_field=True,
            use_color=False,
        )
        captured = capsys.readouterr()
        # Console grouped output was printed
        assert "GROUPED BY FIELD" in captured.out
        # File outputs were also generated
        import os

        generated_files = os.listdir(tmp_path)
        assert any(f.endswith(".json") for f in generated_files)


# ==================== TestPRCommentBreakingChanges ====================


class TestPRCommentBreakingChanges:
    """Test PR comment output breaking changes and truncation edge cases."""

    def test_more_than_10_breaking_changes(self):
        """Line 3524: PR comment truncates breaking changes to 10 with '... more'."""
        from cja_auto_sdr.generator import write_diff_pr_comment_output

        diffs = [
            ComponentDiff(
                id=f"m{i}",
                name=f"Met{i}",
                change_type=ChangeType.MODIFIED,
                changed_fields={"type": (f"old{i}", f"new{i}")},
            )
            for i in range(15)
        ]
        dr = _make_diff_result(
            metric_diffs=diffs,
            summary_kwargs={"metrics_modified": 15},
        )
        output = write_diff_pr_comment_output(dr)
        assert "+5 more" in output

    def test_more_than_25_metric_changes(self):
        """Line 3548: PR comment truncates metric changes to 25 with '... more'."""
        from cja_auto_sdr.generator import write_diff_pr_comment_output

        diffs = [
            ComponentDiff(
                id=f"m{i}",
                name=f"Met{i}",
                change_type=ChangeType.ADDED,
            )
            for i in range(30)
        ]
        dr = _make_diff_result(
            metric_diffs=diffs,
            summary_kwargs={"metrics_added": 30, "metrics_modified": 0},
        )
        output = write_diff_pr_comment_output(dr)
        assert "+5 more" in output

    def test_more_than_25_dimension_changes(self):
        """Line 3566: PR comment truncates dimension changes to 25 with '... more'."""
        from cja_auto_sdr.generator import write_diff_pr_comment_output

        diffs = [
            ComponentDiff(
                id=f"d{i}",
                name=f"Dim{i}",
                change_type=ChangeType.ADDED,
            )
            for i in range(30)
        ]
        dr = _make_diff_result(
            dimension_diffs=diffs,
            summary_kwargs={"dimensions_added": 30, "metrics_modified": 0},
        )
        output = write_diff_pr_comment_output(dr)
        assert "+5 more" in output
