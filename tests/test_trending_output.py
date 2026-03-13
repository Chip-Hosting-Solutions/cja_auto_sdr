"""Tests for trending output across all 6 formats."""

import csv
import json
from datetime import UTC, datetime, timedelta
from pathlib import Path
from unittest.mock import MagicMock

from cja_auto_sdr.org.models import (
    ComponentDistribution,
    ComponentInfo,
    DataViewSummary,
    OrgReportConfig,
    OrgReportResult,
    OrgReportTrending,
    TrendingDelta,
    TrendingSnapshot,
)
from cja_auto_sdr.org.trending import compute_deltas
from cja_auto_sdr.org.writers import (
    _ranked_drift_entries,
    build_org_report_json_data,
    write_org_report_console,
    write_org_report_csv,
    write_org_report_excel,
    write_org_report_html,
    write_org_report_markdown,
)

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _make_trending():
    """Build a minimal OrgReportTrending for testing."""
    return OrgReportTrending(
        snapshots=[
            TrendingSnapshot(
                timestamp="2026-01-01T00:00:00Z",
                data_view_count=10,
                component_count=100,
                core_count=80,
                isolated_count=20,
                high_sim_pair_count=2,
                dv_ids={"dv1", "dv2"},
                dv_names={"dv1": "Legacy DV 1", "dv2": "Test DV 2"},
            ),
            TrendingSnapshot(
                timestamp="2026-02-01T00:00:00Z",
                data_view_count=12,
                component_count=120,
                core_count=95,
                isolated_count=25,
                high_sim_pair_count=3,
                dv_ids={"dv1", "dv2", "dv3"},
                dv_names={"dv1": "Test DV 1", "dv2": "Test DV 2", "dv3": "New DV 3"},
            ),
        ],
        deltas=[
            TrendingDelta(
                from_timestamp="2026-01-01T00:00:00Z",
                to_timestamp="2026-02-01T00:00:00Z",
                data_view_delta=2,
                component_delta=20,
                core_delta=15,
                isolated_delta=5,
                high_sim_pair_delta=1,
            ),
        ],
        drift_scores={"dv1": 0.82, "dv2": 0.15, "dv3": 0.45},
        window_size=2,
    )


def _make_result():
    """Build a minimal OrgReportResult."""
    return OrgReportResult(
        timestamp="2026-02-01T00:00:00Z",
        org_id="test_org",
        parameters=OrgReportConfig(),
        data_view_summaries=[
            DataViewSummary(data_view_id="dv1", data_view_name="Test DV 1", metric_count=50, dimension_count=30),
            DataViewSummary(data_view_id="dv2", data_view_name="Test DV 2", metric_count=20, dimension_count=10),
        ],
        component_index={"m1": ComponentInfo(component_id="m1", component_type="metric", data_views={"dv1", "dv2"})},
        distribution=ComponentDistribution(
            core_metrics=["m1"],
            isolated_metrics=[],
            core_dimensions=[],
            isolated_dimensions=[],
        ),
        similarity_pairs=None,
        recommendations=[],
        duration=1.5,
    )


# ---------------------------------------------------------------------------
# Console
# ---------------------------------------------------------------------------


class TestConsoleWithTrending:
    def test_trending_none_unchanged(self, capsys):
        result = _make_result()
        config = OrgReportConfig()
        write_org_report_console(result, config, quiet=False, trending=None)
        output = capsys.readouterr().out
        assert "TRENDING" not in output

    def test_trending_renders_section(self, capsys):
        result = _make_result()
        config = OrgReportConfig()
        trending = _make_trending()
        write_org_report_console(result, config, quiet=False, trending=trending)
        output = capsys.readouterr().out
        assert "TRENDING" in output
        assert "Data Views" in output
        assert "Period Deltas" in output
        assert "+2" in output
        assert "Components" in output
        assert "Top Drift" in output
        assert "Test DV 1 (dv1)" in output

    def test_trending_quiet_suppressed(self, capsys):
        result = _make_result()
        config = OrgReportConfig()
        write_org_report_console(result, config, quiet=True, trending=_make_trending())
        output = capsys.readouterr().out
        assert "TRENDING" not in output


# ---------------------------------------------------------------------------
# JSON
# ---------------------------------------------------------------------------


class TestJsonWithTrending:
    def test_trending_none_no_key(self):
        result = _make_result()
        data = build_org_report_json_data(result, trending=None)
        assert "trending" not in data

    def test_trending_present_adds_key(self):
        result = _make_result()
        trending = _make_trending()
        data = build_org_report_json_data(result, trending=trending)
        assert "trending" in data
        t = data["trending"]
        assert "snapshots" in t
        assert "deltas" in t
        assert "drift_scores" in t
        assert "drift_details" in t
        assert len(t["snapshots"]) == 2
        assert len(t["deltas"]) == 1
        assert t["drift_scores"]["dv1"] == 0.82
        assert t["drift_details"][0] == {
            "data_view_id": "dv1",
            "data_view_name": "Test DV 1",
            "drift_score": 0.82,
        }

    def test_snapshot_fields(self):
        result = _make_result()
        trending = _make_trending()
        data = build_org_report_json_data(result, trending=trending)
        snap = data["trending"]["snapshots"][0]
        assert snap["timestamp"] == "2026-01-01T00:00:00Z"
        assert snap["data_view_count"] == 10
        assert snap["component_count"] == 100
        assert snap["core_count"] == 80
        assert snap["isolated_count"] == 20
        assert snap["high_sim_pair_count"] == 2

    def test_delta_fields(self):
        result = _make_result()
        trending = _make_trending()
        data = build_org_report_json_data(result, trending=trending)
        delta = data["trending"]["deltas"][0]
        assert delta["data_view_delta"] == 2
        assert delta["component_delta"] == 20

    def test_json_uses_effective_totals_for_manual_id_only_snapshots(self):
        result = _make_result()
        snapshots = [
            TrendingSnapshot(
                timestamp="2026-01-01T00:00:00Z",
                dv_ids={"dv1"},
                dv_names={"dv1": "Legacy DV 1"},
            ),
            TrendingSnapshot(
                timestamp="2026-02-01T00:00:00Z",
                dv_ids={"dv1", "dv2"},
                dv_names={"dv1": "Legacy DV 1", "dv2": "New DV 2"},
            ),
        ]
        trending = OrgReportTrending(
            snapshots=snapshots,
            deltas=compute_deltas(snapshots),
            window_size=2,
        )

        data = build_org_report_json_data(result, trending=trending)

        assert [snapshot["data_view_count"] for snapshot in data["trending"]["snapshots"]] == [1, 2]
        assert data["trending"]["deltas"][0]["data_view_delta"] == 1

    def test_json_roundtrip(self):
        """Trending JSON is serializable and roundtrippable."""
        result = _make_result()
        trending = _make_trending()
        data = build_org_report_json_data(result, trending=trending)
        serialized = json.dumps(data, default=str)
        parsed = json.loads(serialized)
        assert parsed["trending"]["window_size"] == 2

    def test_equal_drift_scores_use_stable_dv_id_tiebreaker(self):
        result = _make_result()
        trending = _make_trending()
        trending.drift_scores = {"dv_b": 0.5, "dv_a": 0.5, "dv_c": 0.4}
        trending.snapshots[0].dv_names.update({"dv_a": "DV A", "dv_b": "DV B", "dv_c": "DV C"})
        trending.snapshots[1].dv_names.update({"dv_a": "DV A", "dv_b": "DV B", "dv_c": "DV C"})

        ranked = _ranked_drift_entries(trending)
        assert [entry["data_view_id"] for entry in ranked] == ["dv_a", "dv_b", "dv_c"]

        data = build_org_report_json_data(result, trending=trending)
        assert list(data["trending"]["drift_scores"]) == ["dv_a", "dv_b", "dv_c"]
        assert [entry["data_view_id"] for entry in data["trending"]["drift_details"]] == ["dv_a", "dv_b", "dv_c"]


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------


class TestExcelWithTrending:
    def test_trending_none_no_worksheet(self, tmp_path):
        result = _make_result()
        logger = MagicMock()
        path = write_org_report_excel(result, tmp_path / "test.xlsx", str(tmp_path), logger, trending=None)
        import openpyxl

        wb = openpyxl.load_workbook(path)
        assert "Trending" not in wb.sheetnames
        wb.close()

    def test_trending_adds_worksheet(self, tmp_path):
        result = _make_result()
        logger = MagicMock()
        trending = _make_trending()
        path = write_org_report_excel(result, tmp_path / "test.xlsx", str(tmp_path), logger, trending=trending)
        import openpyxl

        wb = openpyxl.load_workbook(path)
        assert "Trending" in wb.sheetnames
        ws = wb["Trending"]
        # Should have header row + metric rows
        assert ws.max_row >= 2
        assert ws["A1"].value == "Metric"
        assert ws["B1"].value == "Jan 01"
        assert ws["C1"].value == "Feb 01"
        assert ws["A2"].value == "Data Views"
        assert ws["B2"].value == 10
        assert ws["C2"].value == 12
        assert ws["A8"].value == "Period Deltas"
        assert ws["A9"].value == "Metric"
        assert ws["B9"].value == "Jan 01 -> Feb 01"
        assert ws["A10"].value == "Data Views"
        assert ws["B10"].value == 2
        assert ws["A17"].value == "Drift Scores"
        assert ws["A18"].value == "Data View ID"
        assert ws["B18"].value == "Data View Name"
        assert ws["C18"].value == "Drift Score"
        assert ws["A19"].value == "dv1"
        assert ws["B19"].value == "Test DV 1"
        assert ws["C19"].value == 0.82
        conditional_ranges = {str(rule.sqref) for rule in ws.conditional_formatting}
        assert "B2:C6" in conditional_ranges
        assert "B10:B14" in conditional_ranges
        assert "C19:C21" in conditional_ranges
        wb.close()

    def test_trending_supports_snapshot_columns_beyond_z(self, tmp_path):
        result = _make_result()
        logger = MagicMock()
        start = datetime(2026, 1, 1, tzinfo=UTC)
        snapshots = [
            TrendingSnapshot(
                timestamp=(start + timedelta(days=offset)).isoformat().replace("+00:00", "Z"),
                data_view_count=10 + offset,
                component_count=100 + offset,
                core_count=80 + offset,
                isolated_count=20 + offset,
                high_sim_pair_count=2 + offset,
                dv_ids={"dv1"},
            )
            for offset in range(28)
        ]
        trending = OrgReportTrending(snapshots=snapshots, deltas=[], drift_scores={"dv1": 0.82}, window_size=28)

        path = write_org_report_excel(result, tmp_path / "wide.xlsx", str(tmp_path), logger, trending=trending)
        import openpyxl

        wb = openpyxl.load_workbook(path)
        ws = wb["Trending"]
        assert ws["Z1"].value == "Jan 25"
        assert ws["AA1"].value == "Jan 26"
        assert ws["AC1"].value == "Jan 28"
        conditional_ranges = {str(rule.sqref) for rule in ws.conditional_formatting}
        assert "B2:AC6" in conditional_ranges
        wb.close()

    def test_trending_preserves_duplicate_short_date_labels_in_excel(self, tmp_path):
        result = _make_result()
        logger = MagicMock()
        snapshots = [
            TrendingSnapshot(
                timestamp="2026-01-01T00:00:00Z",
                data_view_count=10,
                component_count=100,
                core_count=80,
                isolated_count=20,
                high_sim_pair_count=2,
            ),
            TrendingSnapshot(
                timestamp="2026-01-01T12:00:00Z",
                data_view_count=11,
                component_count=101,
                core_count=81,
                isolated_count=20,
                high_sim_pair_count=3,
            ),
            TrendingSnapshot(
                timestamp="2027-01-01T00:00:00Z",
                data_view_count=12,
                component_count=102,
                core_count=82,
                isolated_count=20,
                high_sim_pair_count=4,
            ),
        ]
        trending = OrgReportTrending(snapshots=snapshots, deltas=[], drift_scores={}, window_size=3)

        path = write_org_report_excel(
            result, tmp_path / "duplicate_dates.xlsx", str(tmp_path), logger, trending=trending
        )
        import openpyxl

        wb = openpyxl.load_workbook(path)
        ws = wb["Trending"]
        assert [ws.cell(row=1, column=column).value for column in range(2, 5)] == ["Jan 01", "Jan 01", "Jan 01"]
        assert [ws.cell(row=2, column=column).value for column in range(2, 5)] == [10, 11, 12]
        wb.close()


# ---------------------------------------------------------------------------
# Markdown
# ---------------------------------------------------------------------------


class TestMarkdownWithTrending:
    def test_trending_none_unchanged(self, tmp_path):
        result = _make_result()
        logger = MagicMock()
        path = write_org_report_markdown(result, tmp_path / "test.md", str(tmp_path), logger, trending=None)
        content = open(path, encoding="utf-8").read()
        assert "Trending" not in content

    def test_trending_adds_section(self, tmp_path):
        result = _make_result()
        logger = MagicMock()
        trending = _make_trending()
        path = write_org_report_markdown(result, tmp_path / "test.md", str(tmp_path), logger, trending=trending)
        content = open(path, encoding="utf-8").read()
        assert "## Trending" in content
        assert "Data Views" in content
        assert "### Period Deltas" in content
        assert "| Data Views | +2 |" in content
        assert "Drift Scores" in content
        assert "Test DV 1" in content


# ---------------------------------------------------------------------------
# HTML
# ---------------------------------------------------------------------------


class TestHtmlWithTrending:
    def test_trending_none_unchanged(self, tmp_path):
        result = _make_result()
        logger = MagicMock()
        path = write_org_report_html(result, tmp_path / "test.html", str(tmp_path), logger, trending=None)
        content = open(path, encoding="utf-8").read()
        assert "trending" not in content.lower() or "Trending" not in content

    def test_trending_adds_section(self, tmp_path):
        result = _make_result()
        logger = MagicMock()
        trending = _make_trending()
        path = write_org_report_html(result, tmp_path / "test.html", str(tmp_path), logger, trending=trending)
        content = open(path, encoding="utf-8").read()
        assert "Trending" in content
        assert "Period Deltas" in content
        assert ">+2<" in content
        assert "drift" in content.lower()
        assert "Test DV 1" in content


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------


class TestCsvWithTrending:
    def test_trending_none_no_trending_file(self, tmp_path):
        result = _make_result()
        logger = MagicMock()
        write_org_report_csv(result, None, str(tmp_path), logger, trending=None)
        trending_files = list(tmp_path.glob("**/*trending*"))
        assert trending_files == []

    def test_trending_creates_files(self, tmp_path):
        result = _make_result()
        logger = MagicMock()
        trending = _make_trending()
        write_org_report_csv(result, None, str(tmp_path), logger, trending=trending)
        trending_files = list(tmp_path.glob("**/*trending*"))
        assert len(trending_files) >= 1

    def test_trending_csv_rows_match_snapshot_schema(self, tmp_path):
        result = _make_result()
        logger = MagicMock()
        trending = _make_trending()
        csv_dir = Path(write_org_report_csv(result, None, str(tmp_path), logger, trending=trending))

        with (csv_dir / "org_report_trending.csv").open(encoding="utf-8", newline="") as handle:
            rows = list(csv.DictReader(handle))

        assert len(rows) == 10
        assert rows[0] == {
            "Snapshot Timestamp": "2026-01-01T00:00:00Z",
            "Metric": "data_view_count",
            "Value": "10",
        }
        assert rows[-1] == {
            "Snapshot Timestamp": "2026-02-01T00:00:00Z",
            "Metric": "high_sim_pair_count",
            "Value": "3",
        }

    def test_trending_delta_csv_rows_match_delta_schema(self, tmp_path):
        result = _make_result()
        logger = MagicMock()
        trending = _make_trending()
        csv_dir = Path(write_org_report_csv(result, None, str(tmp_path), logger, trending=trending))

        with (csv_dir / "org_report_trending_deltas.csv").open(encoding="utf-8", newline="") as handle:
            rows = list(csv.DictReader(handle))

        assert rows == [
            {
                "From Snapshot Timestamp": "2026-01-01T00:00:00Z",
                "To Snapshot Timestamp": "2026-02-01T00:00:00Z",
                "Period": "Jan 01 -> Feb 01",
                "Metric": "data_view_delta",
                "Metric Label": "Data Views",
                "Value": "2",
            },
            {
                "From Snapshot Timestamp": "2026-01-01T00:00:00Z",
                "To Snapshot Timestamp": "2026-02-01T00:00:00Z",
                "Period": "Jan 01 -> Feb 01",
                "Metric": "component_delta",
                "Metric Label": "Components",
                "Value": "20",
            },
            {
                "From Snapshot Timestamp": "2026-01-01T00:00:00Z",
                "To Snapshot Timestamp": "2026-02-01T00:00:00Z",
                "Period": "Jan 01 -> Feb 01",
                "Metric": "core_delta",
                "Metric Label": "Core",
                "Value": "15",
            },
            {
                "From Snapshot Timestamp": "2026-01-01T00:00:00Z",
                "To Snapshot Timestamp": "2026-02-01T00:00:00Z",
                "Period": "Jan 01 -> Feb 01",
                "Metric": "isolated_delta",
                "Metric Label": "Isolated",
                "Value": "5",
            },
            {
                "From Snapshot Timestamp": "2026-01-01T00:00:00Z",
                "To Snapshot Timestamp": "2026-02-01T00:00:00Z",
                "Period": "Jan 01 -> Feb 01",
                "Metric": "high_sim_pair_delta",
                "Metric Label": "High-Sim Pairs",
                "Value": "1",
            },
        ]

    def test_trending_drift_csv_rows_sorted_descending(self, tmp_path):
        result = _make_result()
        logger = MagicMock()
        trending = _make_trending()
        csv_dir = Path(write_org_report_csv(result, None, str(tmp_path), logger, trending=trending))

        with (csv_dir / "org_report_trending_drift.csv").open(encoding="utf-8", newline="") as handle:
            rows = list(csv.DictReader(handle))

        assert rows == [
            {"Data View ID": "dv1", "Data View Name": "Test DV 1", "Drift Score": "0.82"},
            {"Data View ID": "dv3", "Data View Name": "New DV 3", "Drift Score": "0.45"},
            {"Data View ID": "dv2", "Data View Name": "Test DV 2", "Drift Score": "0.15"},
        ]
